import os
import tempfile
import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows


# Application title and description
st.markdown(
    "<h1 style='text-align: center; color: #4B0082;'>Production Plan Data File Updater</h1>",
    unsafe_allow_html=True
)

st.markdown(
    "<p style='text-align: center; color: #4682B4;'>Upload the files, click 'Update File,' and the updated file will be available for download.</p>",
    unsafe_allow_html=True
)

# File upload section
argo_file = st.file_uploader("Upload the Argo file", type=["xlsx"])
production_plan_file = st.file_uploader("Upload the Production Plan file", type=["xlsx"])

if argo_file and production_plan_file:
    st.success("Both files have been uploaded successfully!")

    # Process files on button click
    if st.button("Update File"):
        try:
            # Read the uploaded Argo file
            argo = pd.ExcelFile(argo_file)
            raw_data = pd.read_excel(argo, sheet_name='SAPUI5 Export', header=0)

            # Filter and process data from Argo
            main_df = raw_data[(raw_data["Division"] == 'PCB') &
                               (raw_data["Plan Product Type"] == 'Tool')]
            main_df['Build Product'] = main_df['Build Product'].replace({
                'AOI FINE HT': 'LUMINA HP', 
                'AOI FINE': 'LUMINA HS',
                'LUMINA HT':'LUMINA HP'
            })
            main_df['Build Complete'] = main_df['Build Complete'].replace({1:'YES',0:'NO'})
            main_df = main_df[~main_df['Build Product'].isin([
                'ULTRA DIMENSION 1000', 'VERIWIDE-A', 'VERIFINE-A', 'DIMENSION 6',
                'VERISMART-A', 'ULTRA VERIFINE-A', 'VERIWIDE', 'ULTRA DIMENSION 800 AOI', 
                'ULTRA DIMENSION 700 AOI', 'APEIRON 800SBS', 'TORNADO', 
                'TITANIUM 900', 'CASTOR TOOL', 'ULTRA PERFIX 500 P',
                'VERISMART','AIM 600','ULTRA DIMENSION LV', 'APEIRON 800XT'
            ])]

            column_to_reformat = ['Build Qtr','Ship Qtr']
            for col in column_to_reformat:
                 main_df[col] =  main_df[col].astype(str)
                 main_df[col + ' - Year'] = '20' + main_df[col].str[2:4]
                 main_df[col + ' - Year'] = main_df[col + ' - Year'].astype(int)
                 main_df[col + ' - Quarter'] = main_df[col].str[5]
                 main_df[col + ' - Year'] = pd.to_numeric(main_df[col + ' - Year'], errors='coerce').fillna(0).astype(int)
                 main_df[col + ' - Quarter'] = pd.to_numeric(main_df[col + ' - Quarter'], errors='coerce').fillna(0).astype(int)

            current_year = datetime.now().year
            current_quarter = (datetime.now().month - 1) // 3 + 1
            # Calculate the end year and quarter for the 8 quarters ahead
            end_year = current_year + (current_quarter + 8) // 4
            end_quarter = (current_quarter + 8) % 4

            # Adjust end_quarter and end_year if end_quarter is 0
            if end_quarter == 0:
               end_quarter = 4
               end_year -= 1

            # Filter:
            # 1. For the current and future quarters within the next 8 quarters.
            # 2. Only systems planned to be built from the current quarter onward.
            main_df = [(((main_df['Build Qtr - Year'] == current_year)&(main_df['Build Qtr - Quarter'] >= current_quarter)) |
            ((main_df['Build Qtr - Year'] > current_year)&(main_df['Build Qtr - Year'] < end_year)) |
            ((main_df['Build Qtr - Year'] == end_year)&(main_df['Build Qtr - Quarter'] <= end_quarter))]

            #Add Revenue column next to MFG column (MFG in the quarter- revenue -Y, if not then -N)
            main_df['MFG_year'] = main_df['MFG Commit Date'].dt.year
            main_df['MFG_quarter'] = (main_df['MFG Commit Date'].dt.month - 1) // 3 + 1
            main_df['Revenue'] = 'N'
            main_df.loc[(main_df['MFG_year'] == current_year) & (main_df['MFG_quarter'] == current_quarter), 'Revenue'] = 'Y'
            
          
            main_df = pd.merge(main_df, product_shortcuts[['Build Product', 'Product']], on='Build Product', how='left')
            columns_to_add = [
                'Opt Start', 'Opt WD', 'Opt End', 
                'Assy Start', 'Assy WD', 'Assy End', 
                'Debug Start', 'Debug WD', 'Debug End', 
                'Int Start', 'Int WD', 'Int End', 
                'Pack Start', 'Pack WD', 'Pack End', 
                'Status','Machine Name','OH PD', 'Flex PD', 'Gripper PD', 'Chamber PD', 
                 'Opt Resource', 'Int Resource', 'Assy Resource', 'Room','Pack Needed'
            ]

            for column in columns_to_add:
                main_df[column] = ""
            main_df = pd.merge(main_df, workdays_df[['Product', 'Opt', 'Ass & Mech','Debug', 'Integration', 'Pack']],
                               on='Product', how='left')
            main_df.fillna('', inplace=True)

            # Convert and clean numeric columns
            main_df['Opt WD'] = np.ceil(pd.to_numeric(main_df['Opt'], errors='coerce')).fillna(0).astype(int)
            main_df['Assy WD'] = np.ceil(pd.to_numeric(main_df['Ass & Mech'], errors='coerce')).fillna(0).astype(int)
            main_df['Debug WD'] = np.ceil(pd.to_numeric(main_df['Debug'], errors='coerce')).fillna(0).astype(int)
            main_df['Int WD'] = np.ceil(pd.to_numeric(main_df['Integration'], errors='coerce')).fillna(0).astype(int)
            main_df['Pack WD'] = np.ceil(pd.to_numeric(main_df['Pack'], errors='coerce')).fillna(0).astype(int)
            main_df.drop(columns=['Opt', 'Ass & Mech', 'Integration','Debug', 'Pack', 'Build Product'], inplace=True)

            # Step 1: Drop rows without Argo ID from the previous plan (just to be safe)
            prev_pp = prev_pp.dropna(subset=['Slot ID/UTID'])

            # Step 2: Define columns that may have changed in the new Argo file
            columns_to_update = ['Build Qtr', 'Argo ID', 'Forecast Product', 'Fab Name',
                         'Product Family', 'Product', 'Build Complete', 'MFG Commit Date','Ship Qtr' ,'Revenue']
            
           # Step 3: Update these columns in prev_pp using values from main_df (based on matching Argo ID)
            main_df = main_df.drop_duplicates(subset='Slot ID/UTID', keep='last')
            for col in columns_to_update:
                prev_pp.loc[prev_pp['Slot ID/UTID'].isin(main_df['Slot ID/UTID']), col] = \
                prev_pp.loc[prev_pp['Slot ID/UTID'].isin(main_df['Slot ID/UTID']), 'Slot ID/UTID'].map(
                    main_df.set_index('Slot ID/UTID')[col]
         )

            # Step 4: Get only the new records from main_df that are not in prev_pp
            new_only = main_df[~main_df['Slot ID/UTID'].isin(prev_pp['Slot ID/UTID'])]
            
            # Step 5: Combine the updated previous plan with the new entries (old first, then new)
            combine_df = pd.concat([prev_pp, new_only], ignore_index=True)

            combine_df = combine_df.drop_duplicates(subset='Slot ID/UTID')
            combine_df = combine_df[['Slot ID/UTID','Argo ID','Build Qtr','Forecast Product', 'Fab Name','Machine Name' , 
                         'Product Family', 'Product', 'Build Complete','Status','Opt Resource','Int Resource','Assy Resource','Room','OH PD','Flex PD','Gripper PD','Chamber PD',
                         'Opt Start', 'Opt WD','Opt End','Assy Start', 'Assy WD', 'Assy End', 'Debug Start', 'Debug WD', 'Debug End', 'Int Start', 'Int WD', 'Int End',
                  'Pack Start', 'Pack WD', 'Pack End', 'Pack Needed', 'MFG Commit Date','Ship Qtr' ,'Revenue']]

            # Load the original workbook and update the data
            wb = load_workbook(production_plan_file)
            ws = wb['Production Plan']

            # Apply styles and update the worksheet
            border_style = Border(left=Side(border_style='thin', color='000000'),
                                  right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
            fill_green = PatternFill(start_color='d9ecd0', end_color='d9ecd0', fill_type='solid')
            fill_gray = PatternFill(start_color='F2F2F2', end_color='D3D3D3', fill_type='solid')
            fill_blue = PatternFill(start_color='B8CCE4', end_color='c0ded9', fill_type='solid')

            def apply_common_style(ws, df):
                start_row = 18
                start_col = 1
                
               #deleting old values
                skip_columns = [19, 21, 22,  24, 25, 27, 28, 30, 31, 33, 34]
                for i in range(start_row, 500):
                    for j in range(1, 38):
                         if j not in skip_columns:
                             ws.cell(row=i, column=j).value = None
                             ws.cell(row=i, column=j).border = None


                for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                    for c_idx, v in enumerate(r, start=start_col):
                        if c_idx in skip_columns:
                            continue
                        cell = ws.cell(row=r_idx, column=c_idx)
                        cell.value = v
                        if r_idx >= start_row and c_idx <= 37:
                            cell.border = border_style
                        cell.font = Font(name='Calibri', size=10)
                        cell.alignment = Alignment(horizontal='center')
                        if r_idx == start_row:
                            if 10 <= c_idx <= 19 or c_idx == 22 or c_idx == 6:
                                cell.fill = fill_gray 
                            else:
                                cell.fill = fill_blue 

            apply_common_style(ws, combine_df)
            #apply current date and time
            current_datetime = datetime.now().strftime("%Y-%m-%d")
           # Write the current date and time to a specific cell, for example, cell AH2
            ws['AH15'] = f"{current_datetime}"


            # Save the file to a temporary directory
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                temp_file_path = tmp_file.name
                wb.save(temp_file_path)

            # Indicate that the file is ready for download
            st.success("The file has been updated successfully and is ready for download!")
            # Get today's date
            today_date = datetime.now().strftime('%Y-%m-%d')

            # Create the new file name with today's date
            new_file_name = f"PCB_GANTT_{today_date}.xlsx"

            # Provide a download link to the user
            with open(temp_file_path, "rb") as file:
                st.download_button(
                    label="Download Updated File",
                    data=file,
                    file_name=new_file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"An error occurred: {e}")
